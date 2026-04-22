from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import openpyxl
import json
import os
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
import io
import re
import pdfplumber
import difflib
import unicodedata

app = Flask(__name__, static_folder='static')
CORS(app)

@app.route('/')
def index():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/<path:path>')
def static_files(path):
    return send_from_directory(app.static_folder, path)

SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

def get_sheets_service():
    creds_json = os.environ.get('GOOGLE_CREDENTIALS')
    if not creds_json:
        raise Exception("Credenciais do Google não configuradas (GOOGLE_CREDENTIALS).")
    creds = Credentials.from_service_account_info(json.loads(creds_json), scopes=SCOPES)
    return build('sheets', 'v4', credentials=creds)

# ===========================================================================
# CONSTANTES
# ===========================================================================

MAPA_SUBCAT = {
    'BEBIDAS NAO ALCOOLICAS':  'BEBIDAS NAO ALCOOLICAS',
    'BEBIDAS NÃO ALCOOLICAS':  'BEBIDAS NAO ALCOOLICAS',
    'BEBIDAS ALCOOLICAS':      'BEBIDAS ALCOOLICAS',
    'DESTILADOS':              'DESTILADOS',
    'DOSES':                   'DOSES & OUTROS',
    'DRINKS':                  'DRINK',
    'COMBOS':                  'COMBOS',
    'OUTROS':                  'DOSES & OUTROS',
}

OVERRIDE_CAT = {
    'GELO SACOLINHA': 'BEBIDAS NAO ALCOOLICAS',
}

CAT_INICIO = {
    'BEBIDAS NAO ALCOOLICAS': 16,
    'BEBIDAS ALCOOLICAS':     32,
    'DESTILADOS':             39,
    'COMBOS':                 52,
    'DRINK':                  68,
    'DOSES & OUTROS':         79,
}

CAT_MAX = {
    'BEBIDAS NAO ALCOOLICAS': 15,
    'BEBIDAS ALCOOLICAS':     9,
    'DESTILADOS':             13,
    'COMBOS':                 16,
    'DRINK':                  11,
    'DOSES & OUTROS':         9,
}

OFFSET_ESTOQUE = -10

OFFSET_PRODUCAO = {
    'BEBIDAS NAO ALCOOLICAS': -11,
    'BEBIDAS ALCOOLICAS':     -10,
    'DESTILADOS':             -10,
    'COMBOS':                 -11,
    'DRINK':                  -11,
    'DOSES & OUTROS':         -10,
}

LIMITE_GARCONS = 30
LIMITE_VOLANTES  = 15

# Variantes de nomes de abas para auto-detecção (Ponto 6)
VARIANTES_ABAS = {
    'CADASTRO':          ['CADASTRO', 'Cadastro', 'cadastro'],
    'ESTOQUE':           ['ESTOQUE', 'Estoque', 'estoque'],
    'PRODUÇÃO':          ['PRODUÇÃO', 'Produção', 'PRODUCAO', 'Producao', 'PRODUC\u00c7\u00c3O'],
    'RESUMO':            ['RESUMO', 'Resumo', 'resumo'],
    'FECHAMENTO CAIXAS': ['FECHAMENTO CAIXAS', 'FECHAMENTO DE CAIXAS', 'Fechamento Caixas',
                          'Fechamento de Caixas', 'FECHAMENTO'],
}

# ===========================================================================
# PONTO 6 — Auto-detectar nomes reais das abas
# ===========================================================================

def detectar_abas(service, spreadsheet_id):
    """
    Lê os nomes reais das abas da planilha e mapeia para os nomes esperados.
    Retorna (dict_mapa, list_nao_encontradas).
    Ex: {'PRODUÇÃO': 'Produção', 'FECHAMENTO CAIXAS': 'FECHAMENTO DE CAIXAS', ...}
    """
    try:
        meta      = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        abas_reais = {s['properties']['title'] for s in meta.get('sheets', [])}
    except Exception:
        abas_reais = set()

    mapa           = {}
    nao_encontradas = []

    for esperado, variantes in VARIANTES_ABAS.items():
        encontrada = next((v for v in variantes if v in abas_reais), None)
        if encontrada:
            mapa[esperado] = encontrada
        else:
            mapa[esperado] = esperado  # fallback: usa o nome esperado
            if abas_reais:
                nao_encontradas.append(f"'{esperado}' (abas disponíveis: {', '.join(sorted(abas_reais))})")

    return mapa, nao_encontradas


def aba(mapa, chave):
    """Helper — retorna nome real da aba ou fallback."""
    return mapa.get(chave, chave)


# ===========================================================================
# PONTO 15 — Validação de arquivos YUZER
# ===========================================================================

def _validar_header_xlsx(ws, campo_esperado, nome_arquivo):
    """Verifica se o xlsx tem o campo esperado nas primeiras 20 linhas."""
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 20:
            break
        if row[0] == campo_esperado:
            return True
    raise ValueError(
        f"'{nome_arquivo}' não parece ser um arquivo válido do YUZER. "
        f"Esperado campo '{campo_esperado}' no cabeçalho. Verifique se enviou o arquivo correto."
    )

# ===========================================================================
# PARSERS
# ===========================================================================

def parse_produtos_xlsx(file_bytes):
    """
    Lê relatório de produtos YUZER.
    Retorna (produtos, is_cortesia, nome_evento).
    Auto-detecta cortesia e extrai nome do evento (Ponto 12).
    Valida que é um arquivo YUZER (Ponto 15).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    is_cortesia  = False
    nome_evento  = None

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 15:
            break
        if row[0]:
            cell = str(row[0])
            if 'Painel de vendas:' in cell:
                nome_evento = cell.replace('Painel de vendas:', '').strip()
                if 'CORTESIA' in nome_evento.upper():
                    is_cortesia = True

    # Ponto 15 — validar que tem header 'Produto'
    _validar_header_xlsx(ws, 'Produto', 'Produtos Vendidos')

    produtos     = []
    header_found = False
    col_map      = {}

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] == 'Produto':
                header_found = True
                col_map = {str(v).strip(): j for j, v in enumerate(row) if v}
            continue

        if row[0] is None:
            continue

        subcat_idx = col_map.get('Subcategoria', 3)
        if subcat_idx >= len(row) or not row[subcat_idx]:
            continue

        subcat = str(row[subcat_idx]).strip().upper()
        cat    = MAPA_SUBCAT.get(subcat, 'DOSES & OUTROS')
        nome   = str(row[0]).strip()

        if nome in OVERRIDE_CAT:
            cat = OVERRIDE_CAT[nome]

        qtd_idx   = col_map.get('Quantidade', 7)
        preco_idx = col_map.get('Preço', 8)

        try:
            qtd = int(float(str(row[qtd_idx] or 0).replace(',', '.')))
        except Exception:
            qtd = 0

        try:
            v = row[preco_idx]
            if isinstance(v, (int, float)):
                preco = round(float(v), 2)
            else:
                s = str(v or 0).replace('R$', '').replace('\xa0', '').strip()
                if ',' in s:
                    s = s.replace('.', '').replace(',', '.')
                preco = round(float(s), 2)
        except Exception:
            preco = 0.0

        if qtd <= 0:
            continue

        produtos.append({
            'produto':      nome,
            'subcategoria': subcat,
            'cat':          cat,
            'qtd_vendida':  qtd,
            'preco':        preco,
        })

    return produtos, is_cortesia, nome_evento


def _preco_str(s):
    if isinstance(s, (int, float)):
        return round(float(s), 2)
    s = str(s or '0').replace('R$', '').replace('\xa0', '').strip()
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return round(float(s), 2)
    except Exception:
        return 0.0


def _normalizar_subcat(s):
    s = str(s).strip().upper().replace('\n', ' ')
    if 'NÃO' in s or 'NAO' in s:
        return 'BEBIDAS NÃO ALCOOLICAS'
    if s in ('BEBIDAS', 'BEBIDAS ALCOOLICAS'):
        return 'BEBIDAS ALCOOLICAS'
    return s


def parse_bonus_pdf(file_bytes):
    """Parser PDF de bônus/cortesia do YUZER."""
    produtos = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                for row in table:
                    if not row or not row[0]:
                        continue
                    cell0 = str(row[0]).strip()
                    if cell0 == 'NOME' or not cell0 or re.match(r'^[\d\s]+$', cell0):
                        continue
                    if row[1] is not None and row[5] is not None:
                        try:
                            qtd = int(str(row[5]).strip())
                            if qtd <= 0:
                                continue
                            subcat = _normalizar_subcat(row[3] or '')
                            cat    = MAPA_SUBCAT.get(subcat, 'DOSES & OUTROS')
                            produtos.append({
                                'produto':     cell0,
                                'cat':         cat,
                                'qtd_vendida': qtd,
                                'preco':       _preco_str(row[8]),
                            })
                        except Exception:
                            pass
                    elif row[1] is None and '\n' in cell0:
                        lines  = cell0.split('\n')
                        subcat = _normalizar_subcat(lines[0])
                        cat    = MAPA_SUBCAT.get(subcat, 'DOSES & OUTROS')
                        for part in lines[1:]:
                            m = re.match(
                                r'^(.+?)\s+FINAL\s+\S+\s+.+?\s+(\d+)\s+\d+\s+\d+\s+R\$\s*([\d.,]+)',
                                part.strip()
                            )
                            if m and int(m.group(2)) > 0:
                                produtos.append({
                                    'produto':     m.group(1).strip(),
                                    'cat':         cat,
                                    'qtd_vendida': int(m.group(2)),
                                    'preco':       round(float(m.group(3).replace('.', '').replace(',', '.')), 2),
                                })
                    elif row[1] is None and 'FINAL' in cell0:
                        m = re.match(
                            r'^(.+?)\s+FINAL\s+\S+\s+(\S+)\s+\S+\s+(\d+)\s+\d+\s+\d+\s+R\$\s*([\d.,]+)',
                            cell0
                        )
                        if m and int(m.group(3)) > 0:
                            subcat = _normalizar_subcat(m.group(2))
                            cat    = MAPA_SUBCAT.get(subcat, 'DOSES & OUTROS')
                            produtos.append({
                                'produto':     m.group(1).strip(),
                                'cat':         cat,
                                'qtd_vendida': int(m.group(3)),
                                'preco':       round(float(m.group(4).replace('.', '').replace(',', '.')), 2),
                            })
    return produtos


def parse_caixas(file_bytes):
    """
    Lê exportação de caixas YUZER.
    Ponto 15 — valida header 'Id'.
    gcol/scol definidos fora do loop (fix anterior).
    op_norm usa unicodedata (fix anterior).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Ponto 15 — validar
    _validar_header_xlsx(ws, 'Id', 'Exportação de Caixas')

    caixas       = []
    col_map      = {}
    header_found = False

    def norm_key(s):
        s = unicodedata.normalize('NFKD', str(s).strip().lower())
        return ''.join(c for c in s if not unicodedata.combining(c))

    def gcol(row, names, fallback):
        for n in names:
            if n in col_map and col_map[n] < len(row):
                try:
                    return round(float(str(row[col_map[n]] or 0).replace(',', '.').replace('R$', '')), 2)
                except Exception:
                    pass
        try:
            return round(float(row[fallback] or 0), 2)
        except Exception:
            return 0.0

    def scol(row, names, fallback):
        for n in names:
            if n in col_map and col_map[n] < len(row):
                return str(row[col_map[n]] or '').strip()
        return str(row[fallback] or '').strip() if fallback < len(row) else ''

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] == 'Id':
                header_found = True
                col_map = {norm_key(v): j for j, v in enumerate(row) if v}
            continue

        if not row[0] or len(str(row[0]).strip()) < 15:
            continue

        dinheiro_bruto = gcol(row, ['dinheiro'], 16)
        devolvido      = gcol(row, ['total produtos retornados', 'total retornado'], 12)

        caixas.append({
            'usuario':  scol(row, ['usuario'], 1),
            'serial':   scol(row, ['serial'], 3),
            'operacao': scol(row, ['operacao'], 5),
            'total':    gcol(row, ['total'], 6),
            'credito':  gcol(row, ['credito'], 13),
            'debito':   gcol(row, ['debito'], 14),
            'pix':      gcol(row, ['pix'], 15),
            'dinheiro': round(max(0.0, dinheiro_bruto - devolvido), 2),
        })

    return caixas


def parse_painel_vendas(file_bytes):
    """
    Lê painel de vendas YUZER.
    Ponto 15 — valida estrutura mínima.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    painel           = {}
    formas           = {}
    lendo_formas     = False
    passou_operacoes = False
    encontrou_total  = False

    FORMAS_PRINCIPAIS = {'PIX', 'DEBIT_CARD', 'CREDIT_CARD', 'CASH', 'APP', 'CASHLESS'}

    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue
        key = str(row[0]).strip()
        val = row[1] if len(row) > 1 else None

        if key == 'Total':
            encontrou_total = True

        if key == 'Formas de Pagamento':
            lendo_formas = True
            continue

        if key.startswith('Total por bandeira') or 'Opera' in key:
            lendo_formas = False
            if 'Opera' in key:
                passou_operacoes = True
            continue

        if lendo_formas and val is not None and key in FORMAS_PRINCIPAIS:
            try:
                formas[key] = round(float(val or 0), 2)
            except Exception:
                pass
            continue

        if not passou_operacoes and key in ('Total', 'Pedidos', 'Média', 'Media', 'Ticket'):
            if key not in painel:
                painel[key] = val

    # Ponto 15 — validar
    if not encontrou_total:
        raise ValueError(
            "Arquivo não reconhecido como Painel de Vendas do YUZER. "
            "Verifique se enviou o arquivo 'exportacao_painel_de_vendas_*.xlsx'."
        )

    painel['formas_pagamento'] = formas
    return painel


# ===========================================================================
# LEITURA DA PLANILHA GOOGLE
# ===========================================================================

def ler_cadastro(service, spreadsheet_id, abas_mapa):
    """
    Lê CADASTRO col B (nome) e F (preço) por categoria.
    Usa nome real da aba (Ponto 6).
    """
    nome_aba = aba(abas_mapa, 'CADASTRO')
    catalogo = {cat: [] for cat in CAT_INICIO}

    for cat, inicio in CAT_INICIO.items():
        fim    = inicio + CAT_MAX[cat] - 1
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"{nome_aba}!B{inicio}:F{fim}"
        ).execute()

        for i, row in enumerate(result.get('values', [])):
            nome_prod = str(row[0]).strip() if len(row) > 0 and row[0] else ''
            if not nome_prod:
                continue
            if nome_prod.upper() in ('BEBIDAS NÃO ALCOOLICAS', 'BEBIDAS ALCOOLICAS',
                                      'DESTILADOS', 'COMBOS', 'DRINK', 'DOSES & OUTROS',
                                      'VALOR', 'CARDAPIO', 'BRIEFING'):
                continue
            try:
                raw   = row[4] if len(row) > 4 else 0
                preco = round(float(raw), 2) if isinstance(raw, (int, float)) else \
                        round(float(str(raw).replace('R$', '').replace(',', '.').strip()), 2)
            except Exception:
                preco = 0.0
            if preco <= 0:
                continue
            catalogo[cat].append({
                'nome':           nome_prod,
                'preco':          preco,
                'linha_cadastro': inicio + i,
            })

    return catalogo


def ler_mapa_linhas(service, spreadsheet_id, abas_mapa):
    """
    Lê col A de ESTOQUE e PRODUÇÃO → nome→linha.
    Ponto 6 — usa nomes reais das abas.
    Ponto 5 (anterior) — tenta múltiplas variações de PRODUÇÃO.
    """
    IGNORAR = {
        'PRODUTO', 'BEBIDAS NÃO ALCOOLICAS', 'BEBIDAS ALCOOLICAS', 'DESTILADOS',
        'COMBOS', 'DRINK', 'DOSES & OUTROS', 'FECHAMENTO GERAL BAR CONSUMO/VENDA',
        'OBSERVAÇÃO PREENCHER APENAS AS COLUNAS EM AMARELO',
        'CONSUMO PRODUÇÃO CAMARIM / BONUS', 'RESUMO ALIMENTAÇAO',
        'TOTAL / CARTÃO', 'TOTAL',
    }
    est_map  = {}
    prod_map = {}

    nome_est = aba(abas_mapa, 'ESTOQUE')
    r = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id, range=f"{nome_est}!A1:A80"
    ).execute()
    for i, row in enumerate(r.get('values', []), 1):
        if row and row[0] and str(row[0]).strip() not in IGNORAR:
            est_map[str(row[0]).strip()] = i

    # Tentar múltiplas variações do nome da aba PRODUÇÃO
    nome_prod_aba = aba(abas_mapa, 'PRODUÇÃO')
    variantes_prod = [nome_prod_aba] + [v for v in VARIANTES_ABAS.get('PRODUÇÃO', [])
                                         if v != nome_prod_aba]
    for nome_aba_p in variantes_prod:
        try:
            rp = service.spreadsheets().values().get(
                spreadsheetId=spreadsheet_id, range=f"{nome_aba_p}!A1:A80"
            ).execute()
            for i, row in enumerate(rp.get('values', []), 1):
                if row and row[0] and str(row[0]).strip() not in IGNORAR:
                    prod_map[str(row[0]).strip()] = i
            if prod_map:
                break
        except Exception:
            continue

    return est_map, prod_map


# ===========================================================================
# MOTOR DE CONCILIAÇÃO MULTI-ATRIBUTO
# ===========================================================================

STOP_WORDS = {
    'ml', 'l', 'lt', 'kg', 'g', 'un', 'und', 'cx',
    '100', '130', '150', '160', '180', '200', '220', '250', '260', '269',
    '280', '290', '300', '320', '330', '340', '350', '355', '360', '370',
    '380', '410', '430', '440', '473', '500', '550', '600', '750', '1000',
    'com', 'de', 'do', 'da', 'e', 'em', 'para', 'o', 'a', 'os', 'as',
    'final', 'bebidas', 'drink', 'dose', 'doses', 'combo', 'outros',
    'garrafa', 'lata', 'long', 'neck', 'anos',
    '1', '2', '3', '4', '5', '6', '7', '8', '9', '10',
}

ALIAS = {
    'buul': 'bull', 'redbull': 'red bull', 'redbuul': 'red bull',
    'oldpar': 'old parr', 'oldparr': 'old parr',
    'tanquery': 'tanqueray', 'mulle': 'mule',
    'limonade': 'lemonade', 'tonica': 'tonica', 'tonicas': 'tonica',
}


def _norm_str(s):
    s = unicodedata.normalize('NFKD', str(s))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'[^a-z0-9 ]', ' ', s.lower())
    for a_key, a_val in ALIAS.items():
        s = re.sub(r'\b' + a_key + r'\b', a_val, s)
    return s


def _tokens(nome):
    return set(t for t in _norm_str(nome).split() if t not in STOP_WORDS and len(t) > 1)


def _extrair_ml(nome):
    m = re.search(r'(\d+\.?\d*)\s*(ml|l|lt)\b', nome.lower().replace(' ', ''))
    if m:
        val = float(m.group(1))
        return val * 1000 if m.group(2) in ('l', 'lt') else val
    return None


def _score_par(nome_p, preco_p, nome_y, preco_y):
    """Score: 35% preço + 45% nome (seq+tokens) + 20% unidade."""
    if preco_p > 0 and preco_y > 0:
        s_preco = 1.0 if abs(preco_p - preco_y) < 0.01 else \
                  max(0.0, 1.0 - abs(preco_p - preco_y) / max(preco_p, preco_y) * 3)
    else:
        s_preco = 0.0

    s_seq  = difflib.SequenceMatcher(None, _norm_str(nome_p), _norm_str(nome_y)).ratio()
    tp, ty = _tokens(nome_p), _tokens(nome_y)
    s_tok  = len(tp & ty) / len(tp | ty) if (tp and ty) else 0.0
    s_nome = s_seq * 0.6 + s_tok * 0.4

    ml_p, ml_y = _extrair_ml(nome_p), _extrair_ml(nome_y)
    s_un = 1.0 if (ml_p and ml_y and abs(ml_p - ml_y) < 1) else 0.5

    return s_preco * 0.35 + s_nome * 0.45 + s_un * 0.20


LIMIAR_SCORE = 0.40


def conciliar(catalogo, vendas, bonus, mapeamento=None):
    """Pré-match global greedy por score descendente por categoria."""
    mapeamento = mapeamento or {}
    agrupado   = {cat: [] for cat in CAT_INICIO}

    def aplicar_mapa(lista):
        resultado = []
        for p in lista:
            if p['produto'] in mapeamento:
                p = dict(p)
                p['produto'] = mapeamento[p['produto']]
            resultado.append(p)
        return resultado

    vendas_map = aplicar_mapa(vendas)
    bonus_map  = aplicar_mapa(bonus)

    def pre_match(planilha_items, yuzer_items):
        if not planilha_items or not yuzer_items:
            return {}
        pares = []
        for pi, item in enumerate(planilha_items):
            for yi, prod in enumerate(yuzer_items):
                s = _score_par(item['nome'], item['preco'], prod['produto'], prod['preco'])
                if s >= LIMIAR_SCORE:
                    pares.append((s, pi, yi))
        pares.sort(reverse=True)
        plan_usado  = set()
        yuzer_usado = set()
        resultado   = {}
        for s, pi, yi in pares:
            if pi not in plan_usado and yi not in yuzer_usado:
                resultado[pi] = (yi, yuzer_items[yi]['produto'], s, yuzer_items[yi]['qtd_vendida'])
                plan_usado.add(pi)
                yuzer_usado.add(yi)
        return resultado

    for cat, itens in catalogo.items():
        v_cat   = [p for p in vendas_map if p['cat'] == cat]
        b_cat   = [p for p in bonus_map  if p['cat'] == cat]
        match_v = pre_match(itens, v_cat)
        match_b = pre_match(itens, b_cat)

        for pi, item in enumerate(itens):
            mv = match_v.get(pi)
            mb = match_b.get(pi)
            agrupado[cat].append({
                'nome':           item['nome'],
                'linha_cadastro': item['linha_cadastro'],
                'preco':          item['preco'],
                'qtd_venda':      mv[3] if mv else 0,
                'qtd_bonus':      mb[3] if mb else 0,
                'qtd_sistema':    (mv[3] if mv else 0) + (mb[3] if mb else 0),
                'match_venda':    mv[1] if mv else None,
                'score_venda':    mv[2] if mv else 0.0,
                'match_bonus':    mb[1] if mb else None,
                'score_bonus':    mb[2] if mb else 0.0,
                'conciliado':     mv is not None or mb is not None,
            })

    return agrupado


def gerar_sugestoes_mapeamento(vendas, bonus, catalogo, agrupado):
    """Sugestões apenas para produtos YUZER sem match (fix anterior)."""
    ja_matchados = set()
    for prods in agrupado.values():
        for p in prods:
            if p.get('match_venda'): ja_matchados.add(p['match_venda'])
            if p.get('match_bonus'): ja_matchados.add(p['match_bonus'])

    todos_planilha = [p for prods in catalogo.values() for p in prods]
    sugestoes      = []
    vistos         = set()

    for yu in (vendas + bonus):
        nome_yu = yu['produto']
        if nome_yu in ja_matchados or nome_yu in vistos:
            continue
        vistos.add(nome_yu)
        scores = sorted(
            [(s, pl['nome'], pl['preco'])
             for pl in todos_planilha
             for s in [_score_par(pl['nome'], pl['preco'], nome_yu, yu['preco'])]
             if s >= 0.30],
            reverse=True
        )
        sugestoes.append({
            'yuzer':        nome_yu,
            'preco_yuzer':  yu['preco'],
            'cat':          yu['cat'],
            'qtd':          yu['qtd_vendida'],
            'sugestao':     scores[0][1] if scores else None,
            'score':        round(scores[0][0], 2) if scores else 0,
            'alternativas': [{'nome': s[1], 'preco': s[2], 'score': round(s[0], 2)}
                             for s in scores[1:4]],
        })

    return sugestoes


# ===========================================================================
# PONTO 13 — Tabela de destino célula a célula
# ===========================================================================

def build_destino_preview(agrupado, est_map, prod_map, abas_mapa, painel_data=None):
    """
    Monta tabela auditável mostrando exatamente quais células serão preenchidas.
    Retorna dict com 'estoque', 'producao', 'resumo'.
    """
    nome_est  = aba(abas_mapa, 'ESTOQUE')
    nome_prod = aba(abas_mapa, 'PRODUÇÃO')
    nome_res  = aba(abas_mapa, 'RESUMO')

    destino_estoque  = []
    destino_producao = []

    for cat, prods in agrupado.items():
        for p in prods:
            if p['qtd_venda'] > 0:
                linha = est_map.get(p['nome']) or (p['linha_cadastro'] + OFFSET_ESTOQUE)
                destino_estoque.append({
                    'nome':   p['nome'],
                    'celula': f"{nome_est}!I{linha}",
                    'valor':  p['qtd_venda'],
                    'match':  p.get('match_venda', ''),
                    'score':  round(p.get('score_venda', 0), 2),
                })

            if p['qtd_bonus'] > 0:
                linha = prod_map.get(p['nome']) or (p['linha_cadastro'] + OFFSET_PRODUCAO.get(cat, -10))
                destino_producao.append({
                    'nome':   p['nome'],
                    'celula': f"{nome_prod}!C{linha}",
                    'valor':  p['qtd_bonus'],
                })

    # RESUMO B3:B9
    destino_resumo = []
    if painel_data:
        fp = painel_data.get('formas_pagamento', {})
        total_fp = sum(fp.values())
        for linha, label, val in [
            (3,  'APP',          0),
            (4,  'Dinheiro',     fp.get('CASH', 0)),
            (5,  'Crédito',      fp.get('CREDIT_CARD', 0)),
            (6,  'Débito',       fp.get('DEBIT_CARD', 0)),
            (7,  'PIX',          fp.get('PIX', 0)),
            (8,  'Cancelamento', 0),
            (9,  'Receita Total', total_fp),
        ]:
            destino_resumo.append({
                'celula': f"{nome_res}!B{linha}",
                'label':  label,
                'valor':  val,
            })

    return {
        'estoque':  destino_estoque,
        'producao': destino_producao,
        'resumo':   destino_resumo,
    }


# ===========================================================================
# BUILDERS GOOGLE SHEETS
# ===========================================================================

def build_estoque_updates(agrupado, est_map, abas_mapa):
    """ESTOQUE col I = qtd_venda. Não escreve zeros (fix anterior)."""
    nome_est = aba(abas_mapa, 'ESTOQUE')
    updates  = []
    for prods in agrupado.values():
        for p in prods:
            if p['qtd_venda'] <= 0:
                continue
            nome  = p['nome']
            linha = est_map.get(nome) or (p['linha_cadastro'] + OFFSET_ESTOQUE)
            if linha and linha > 0:
                updates.append({'range': f"{nome_est}!I{linha}", 'values': [[p['qtd_venda']]]})
    return updates


def build_producao_updates(agrupado, prod_map, abas_mapa):
    """PRODUÇÃO col C = qtd_bonus."""
    nome_prod = aba(abas_mapa, 'PRODUÇÃO')
    updates   = []
    avisos    = []
    for cat, prods in agrupado.items():
        for p in prods:
            if p['qtd_bonus'] <= 0:
                continue
            nome  = p['nome']
            linha = prod_map.get(nome) or (p['linha_cadastro'] + OFFSET_PRODUCAO.get(cat, -10))
            if linha and linha > 0:
                updates.append({'range': f"{nome_prod}!C{linha}", 'values': [[p['qtd_bonus']]]})
            if p.get('score_bonus', 1.0) < 0.40 and p.get('match_bonus'):
                avisos.append(f"Bônus incerto: '{nome}' → '{p['match_bonus']}' ({int(p['score_bonus']*100)}%)")
    return updates, avisos


# ===========================================================================
# MAPEAMENTO (memória + persistência via env var)
# ===========================================================================

def _carregar_mapeamento_persistido():
    raw = os.environ.get('PRIMEBAR_MAPEAMENTO', '')
    if raw:
        try:
            return json.loads(raw)
        except Exception:
            pass
    return {}


_mapeamento_global = {
    # ── Drinks padrão Prime Bar ──────────────────────────────────────────
    'DRINK Tropical Gin':        'TROPICAL GIN ( GIN + RODELA DE LARANJA E RED BUUL TROPICAL )',
    'DRINK Melancita':           'MELANCITA ( GIN + RODELA DE LIMÃO E RED BUUL MELANCIA )',
    'DRINK Moscow Mule':         'MOSCOW MULLE ( VODKA + XAROPE DE GENGIBRE + SUMO DE LIMÃO E ESPUMA CITRICA )',
    'DRINK Pink Limonade':       'PINK LEMONADE ( GIN +  SUCO DE  LIMÃO + GROSELHA E RODELA DE LIMÃO SICILIANO)',
    'DRINK Gija':                'GIJA ( GIN + TONICA + XAROPE DE GENGIBRE + CANELA E RODELA DE LIMÃO SICILIANO )',
    'DRINK Gin Tônica':          'GIN TONICA ( GIN + TONICA E RODELA DE LIMÃO )',

    # ── Fix 1a: nomes corrigidos para bater com planilha real ────────────
    # (antes tinham typos: RED BUUL, REDBULL sem espaço)
    'DRINK Vodka + Red Bull':    'VODKA E REDBULL ( VODKA + REDBULL',
    'Old Parr+3 Red Bull':       'OLDPAR 12 ANOS 1L + 3 RED BULL 250ML',
    'Old Parr+5 Águas de Coco':  'OLDPARR 12 ANOS 1L + 5 AGUA DE COCO',
    'DRINK Gin + Red Bull':      'GIN COM REDBULL  (GIN + REDBULL',

    # ── Fix 1b: drinks Prime (YUZER abreviado → planilha nome completo) ──
    'DRINK Prime Tonic':         'PRIME TONIC ( GIN  – AGUA TONICA - LIMAO SICILIANO – ALLECRIM)',
    'DRINK Prime Mule':          'PRIME MULE ( VODKA - LIMAO - XAROPE GENGIBRE - AGUA COM GÁS - ESPUMA GENGIBRE)',
    'DRINK Prime Penicillin':    'PRIME PENICILLIN ( WHISK 12 ANOS - SUCO LIMAO SICILIANO, XAROPE GENGIBRE - XAROPE DE AÇÚCAR)',
}
_mapeamento_global.update(_carregar_mapeamento_persistido())
_mapeamento_store = {}


def get_mapa(spreadsheet_id):
    mapa = dict(_mapeamento_global)
    mapa.update(_mapeamento_store.get(spreadsheet_id, {}))
    return mapa


# ===========================================================================
# LIMPEZA DA PLANILHA
# ===========================================================================

def limpar_planilha(service, spreadsheet_id, abas_mapa):
    """
    Zera apenas campos automáticos. Usa nomes reais das abas (Ponto 6).
    Nunca toca em RELATORIO DE VENDA, col D do RESUMO, linhas de total.
    """
    nome_est  = aba(abas_mapa, 'ESTOQUE')
    nome_res  = aba(abas_mapa, 'RESUMO')
    nome_cx   = aba(abas_mapa, 'FECHAMENTO CAIXAS')

    ranges_base = [
        f'{nome_res}!B3:B9',
        f'{nome_est}!I6:I76',
        f'{nome_cx}!B3:H32',    # Garçons PIX
        f'{nome_cx}!B36:H50',   # Caixas Volantes
        f'{nome_cx}!B54:H83',   # Caixas Fixos/PDV
    ]
    service.spreadsheets().values().batchClear(
        spreadsheetId=spreadsheet_id,
        body={'ranges': ranges_base}
    ).execute()

    # Limpar PRODUÇÃO col C — tentar múltiplas variações do nome
    nome_prod_aba = aba(abas_mapa, 'PRODUÇÃO')
    variantes_prod = [nome_prod_aba] + [v for v in VARIANTES_ABAS.get('PRODUÇÃO', [])
                                         if v != nome_prod_aba]
    for nome_aba_p in variantes_prod:
        try:
            service.spreadsheets().values().clear(
                spreadsheetId=spreadsheet_id,
                range=f"{nome_aba_p}!C5:C77"
            ).execute()
            break
        except Exception:
            continue


# ===========================================================================
# VALIDAÇÃO DE TOTAIS
# ===========================================================================

def validar_totais(agrupado, painel):
    """Compara totais conciliados vs painel. Tolerância R$50 (fix anterior)."""
    avisos       = []
    total_painel = float(painel.get('Total', 0) or 0)
    if total_painel == 0:
        return avisos

    total_calc = sum(
        p['qtd_venda'] * p['preco']
        for prods in agrupado.values()
        for p in prods
        if p['qtd_venda'] > 0
    )
    if total_calc == 0:
        return avisos

    diferenca = abs(total_painel - total_calc)
    pct       = diferenca / total_painel * 100

    if diferenca > 50:
        avisos.append(
            f"Divergência: Painel=R${total_painel:,.2f} | "
            f"Conciliado=R${total_calc:,.2f} | Δ=R${diferenca:,.2f} ({pct:.1f}%) "
            f"— pode haver produtos sem slot no CADASTRO"
        )
    return avisos


# ===========================================================================
# RELATÓRIO DE FECHAMENTO
# ===========================================================================

def gerar_relatorio_texto(agrupado, msgs, avisos, painel_data):
    fp  = painel_data.get('formas_pagamento', {})
    tot = painel_data.get('Total', 0)
    linhas = [
        "PRIME BAR — RELATÓRIO DE FECHAMENTO",
        "=" * 50, "",
        "PAGAMENTOS",
        f"  Total:    R${float(tot or 0):>12,.2f}",
        f"  Dinheiro: R${fp.get('CASH', 0):>12,.2f}",
        f"  Crédito:  R${fp.get('CREDIT_CARD', 0):>12,.2f}",
        f"  Débito:   R${fp.get('DEBIT_CARD', 0):>12,.2f}",
        f"  PIX:      R${fp.get('PIX', 0):>12,.2f}",
        "",
        "ESTOQUE col I — produtos vendidos",
        f"  {'Produto':<40} {'Qtd':>5} {'Score':>5}",
        "  " + "-" * 52,
    ]

    # Totais por categoria para o relatório (Ponto 20 — bônus)
    totais_cat = {}
    for cat, prods in agrupado.items():
        qtd_cat = sum(p['qtd_venda'] for p in prods if p['qtd_venda'] > 0)
        rec_cat = sum(p['qtd_venda'] * p['preco'] for p in prods if p['qtd_venda'] > 0)
        if qtd_cat > 0:
            totais_cat[cat] = (qtd_cat, rec_cat)
        for p in prods:
            if p['qtd_venda'] > 0:
                flag = ' ⚠' if p.get('score_venda', 1) < 0.5 else ''
                linhas.append(f"  {p['nome'][:39]:<40} {p['qtd_venda']:>5} {p.get('score_venda',0):>5.2f}{flag}")

    linhas += ["", "TOTAIS POR CATEGORIA"]
    for cat, (qtd, rec) in totais_cat.items():
        linhas.append(f"  {cat:<30} qtd:{qtd:>5}  R${rec:>10,.2f}")

    linhas += ["", "PRODUÇÃO col C — bônus/cortesia"]
    bon = [p for prods in agrupado.values() for p in prods if p['qtd_bonus'] > 0]
    for p in bon:
        linhas.append(f"  {p['nome'][:39]:<40} {p['qtd_bonus']:>5}")
    if not bon:
        linhas.append("  (nenhum)")

    nconc = [p for prods in agrupado.values() for p in prods
             if not p.get('conciliado') and p['preco'] > 0]
    if nconc:
        linhas += ["", "NÃO VENDIDOS NO EVENTO (zerado — esperado)"]
        for p in nconc:
            linhas.append(f"  {p['nome']}")

    if avisos:
        linhas += ["", "AVISOS"]
        for a in avisos:
            linhas.append(f"  ⚠ {a}")

    if msgs:
        linhas += ["", "LOG"]
        for m in msgs:
            linhas.append(f"  ✓ {m}")

    return "\n".join(linhas)


# ===========================================================================
# UTILITÁRIO
# ===========================================================================

def _extrair_sid(raw):
    if not raw:
        return ''
    raw = raw.strip()
    if 'docs.google.com' in raw:
        m = re.search(r'/d/([a-zA-Z0-9-_]+)', raw)
        if m:
            return m.group(1)
    return raw


# ===========================================================================
# ROTAS
# ===========================================================================

@app.route('/api/preview', methods=['POST'])
def preview():
    try:
        result = {}
        vendas = []
        bonus  = []
        nomes_evento = []

        # Múltiplos arquivos — auto-detecta cortesia, extrai nome do evento
        for f in request.files.getlist('produtos_vendidos'):
            raw = f.read()
            try:
                prods, is_cort, nome_ev = parse_produtos_xlsx(raw)
                if nome_ev and nome_ev not in nomes_evento:
                    nomes_evento.append(nome_ev)
                if is_cort:
                    bonus += prods
                else:
                    vendas += prods
            except ValueError as e:
                return jsonify({'success': False, 'error': str(e)}), 400

        result['produtos'] = vendas
        result['bonus']    = bonus  # sempre atribuído

        # Arquivo de bônus explícito
        if 'produtos_bonus' in request.files:
            b = request.files['produtos_bonus']
            try:
                if b.filename.lower().endswith('.pdf'):
                    bonus += parse_bonus_pdf(b.read())
                else:
                    prods, _, nome_ev = parse_produtos_xlsx(b.read())
                    if nome_ev and nome_ev not in nomes_evento:
                        nomes_evento.append(nome_ev)
                    bonus += prods
                result['bonus'] = bonus
            except ValueError as e:
                return jsonify({'success': False, 'error': str(e)}), 400

        # Ponto 12 — nome do evento para confirmação visual
        result['evento'] = nomes_evento[0] if nomes_evento else None

        if 'exportacao_caixas' in request.files:
            try:
                result['caixas'] = parse_caixas(request.files['exportacao_caixas'].read())
            except ValueError as e:
                return jsonify({'success': False, 'error': str(e)}), 400

        if 'painel_de_vendas' in request.files:
            try:
                painel = parse_painel_vendas(request.files['painel_de_vendas'].read())
                fp     = painel.get('formas_pagamento', {})
                result['painel'] = painel
                result['resumo'] = {
                    'total_faturado': painel.get('Total', 0),
                    'total_pedidos':  painel.get('Pedidos', 0),
                    'ticket_medio':   painel.get('Média', painel.get('Media', 0)),
                    'credito':        fp.get('CREDIT_CARD', 0),
                    'debito':         fp.get('DEBIT_CARD', 0),
                    'pix':            fp.get('PIX', 0),
                    'dinheiro':       fp.get('CASH', 0),
                }
            except ValueError as e:
                return jsonify({'success': False, 'error': str(e)}), 400

        sid = _extrair_sid(request.form.get('spreadsheet_id', ''))
        if sid and (vendas or bonus):
            try:
                service  = get_sheets_service()

                # Ponto 6 — detectar abas reais
                abas_mapa, abas_nf = detectar_abas(service, sid)

                catalogo   = ler_cadastro(service, sid, abas_mapa)
                mapa_atual = get_mapa(sid)
                agrupado   = conciliar(catalogo, vendas, bonus, mapa_atual)

                # Sugestões apenas para sem match
                sugestoes = gerar_sugestoes_mapeamento(vendas, bonus, catalogo, agrupado)
                result['sugestoes_mapeamento'] = sugestoes

                # Ponto 13 — tabela destino célula a célula
                try:
                    est_map_p, prod_map_p = ler_mapa_linhas(service, sid, abas_mapa)
                    painel_d = result.get('painel')
                    result['destino'] = build_destino_preview(
                        agrupado, est_map_p, prod_map_p, abas_mapa, painel_d
                    )
                except Exception:
                    pass  # destino é opcional no preview

                # Fix 3: todos_matched contém nomes da planilha, mas vendas tem
                # nomes originais do YUZER. Verificar nome original E nome mapeado
                # para não gerar falsos positivos em produtos que já foram matchados.
                todos_matched = set()
                for ps in agrupado.values():
                    for p in ps:
                        if p.get('match_venda'): todos_matched.add(p['match_venda'])
                        if p.get('match_bonus'):  todos_matched.add(p['match_bonus'])

                yuzer_sem_slot = []
                vistos_ss      = set()
                for p in (vendas + bonus):
                    nome_original = p['produto']
                    nome_mapeado  = mapa_atual.get(nome_original, nome_original)
                    ja_match = (nome_original in todos_matched or
                                nome_mapeado  in todos_matched)
                    if not ja_match and nome_original not in vistos_ss:
                        vistos_ss.add(nome_original)
                        yuzer_sem_slot.append({
                            'produto':     nome_original,
                            'preco':       p['preco'],
                            'qtd_vendida': p['qtd_vendida'],
                            'cat':         p['cat'],
                        })

                result['yuzer_sem_slot'] = yuzer_sem_slot

                total_prod  = sum(len(v) for v in agrupado.values())
                conciliados = sum(1 for ps in agrupado.values() for p in ps if p.get('conciliado'))
                sem_match   = [p for ps in agrupado.values() for p in ps
                               if not p.get('conciliado') and p['preco'] > 0]

                avisos_pv = []
                if abas_nf:
                    avisos_pv.append(f"Abas não encontradas: {'; '.join(abas_nf)}")

                result['validacao'] = {
                    'total_produtos':  total_prod,
                    'conciliados':     conciliados,
                    'sem_match':       [{'nome': p['nome'], 'preco': p['preco']} for p in sem_match],
                    'score_baixo':     [{'planilha': p['nome'], 'yuzer': p['match_venda'],
                                         'score': p['score_venda']}
                                        for ps in agrupado.values() for p in ps
                                        if p.get('match_venda') and 0 < p['score_venda'] < 0.50],
                    'yuzer_sem_slot':  len(yuzer_sem_slot),
                    'abas_detectadas': abas_mapa,
                    'avisos':          avisos_pv,
                    'pode_enviar':     True,
                }

            except Exception as e_prev:
                result['validacao'] = {'erro_preview': str(e_prev), 'pode_enviar': True}

        return jsonify({'success': True, 'data': result})

    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e),
                        'trace': traceback.format_exc()}), 400


@app.route('/api/mapeamento', methods=['GET'])
def get_mapeamento():
    sid  = _extrair_sid(request.args.get('spreadsheet_id', ''))
    mapa = get_mapa(sid)
    return jsonify({
        'success':    True,
        'mapeamento': mapa,
        'global':     _mapeamento_global,
        'especifico': _mapeamento_store.get(sid, {}),
    })


@app.route('/api/mapeamento', methods=['POST'])
def save_mapeamento():
    data = request.get_json(force=True) or {}
    sid  = _extrair_sid(data.get('spreadsheet_id', 'global'))
    mapa = data.get('mapeamento', {})

    if not sid or sid == 'global':
        _mapeamento_global.update(mapa)
    else:
        _mapeamento_store.setdefault(sid, {}).update(mapa)

    if data.get('aprender', False):
        _mapeamento_global.update(mapa)

    return jsonify({
        'success':          True,
        'total_global':     len(_mapeamento_global),
        'total_especifico': len(_mapeamento_store.get(sid, {})),
    })


@app.route('/api/enviar', methods=['POST'])
def enviar():
    try:
        sid = _extrair_sid(request.form.get('spreadsheet_id', ''))
        if not sid:
            return jsonify({'success': False, 'error': 'ID da planilha não informado.'}), 400

        service     = get_sheets_service()
        batch       = []
        msgs        = []
        avisos      = []
        agrupado    = {}
        painel_data = {}

        # Ponto 6 — detectar abas reais antes de qualquer operação
        abas_mapa, abas_nf = detectar_abas(service, sid)
        if abas_nf:
            avisos.append(f"Abas com nome diferente do esperado: {'; '.join(abas_nf)}")
        msgs.append(f"Abas: {', '.join(f'{k}→{v}' for k,v in abas_mapa.items() if k != v) or 'OK (nomes padrão)'}")

        # Limpeza
        try:
            limpar_planilha(service, sid, abas_mapa)
            msgs.append('Campos automáticos zerados')
        except Exception as e_limpa:
            avisos.append(f'Limpeza parcial: {str(e_limpa)[:80]}')

        # ── Produtos ──────────────────────────────────────────────────────
        vendas_files = request.files.getlist('produtos_vendidos')
        if vendas_files:
            vendas       = []
            bonus        = []
            nomes_evento = []

            for f in vendas_files:
                try:
                    prods, is_cort, nome_ev = parse_produtos_xlsx(f.read())
                    if nome_ev and nome_ev not in nomes_evento:
                        nomes_evento.append(nome_ev)
                    if is_cort:
                        bonus += prods
                        msgs.append(f'Cortesia detectada: {len(prods)} produtos → PRODUÇÃO col C')
                    else:
                        vendas += prods
                except ValueError as e:
                    return jsonify({'success': False, 'error': str(e)}), 400

            if nomes_evento:
                msgs.append(f'Evento: {" / ".join(nomes_evento)}')

            if 'produtos_bonus' in request.files:
                b = request.files['produtos_bonus']
                try:
                    if b.filename.lower().endswith('.pdf'):
                        bonus += parse_bonus_pdf(b.read())
                    else:
                        prods, _, _ = parse_produtos_xlsx(b.read())
                        bonus += prods
                except ValueError as e:
                    return jsonify({'success': False, 'error': str(e)}), 400

            mapa_atual = get_mapa(sid)
            catalogo   = ler_cadastro(service, sid, abas_mapa)
            msgs.append(f'CADASTRO: {sum(len(v) for v in catalogo.values())} produtos')

            agrupado    = conciliar(catalogo, vendas, bonus, mapa_atual)
            total_prod  = sum(len(v) for v in agrupado.values())
            conciliados = sum(1 for ps in agrupado.values() for p in ps if p.get('conciliado'))
            nao_vend    = sum(1 for ps in agrupado.values() for p in ps
                              if not p.get('conciliado') and p['preco'] > 0)
            msgs.append(f'Conciliação: {conciliados}/{total_prod}')

            if nao_vend:
                nomes_nv = [p['nome'] for ps in agrupado.values() for p in ps
                            if not p.get('conciliado') and p['preco'] > 0]
                msgs.append(f'ℹ️ {nao_vend} não vendidos neste evento: {"; ".join(nomes_nv[:5])}')

            # Produtos YUZER sem slot
            # Fix 3: verificar nome original E mapeado para evitar falsos positivos
            todos_matched = set()
            for ps in agrupado.values():
                for p in ps:
                    if p.get('match_venda'): todos_matched.add(p['match_venda'])
                    if p.get('match_bonus'):  todos_matched.add(p['match_bonus'])
            sem_slot = list(dict.fromkeys(
                p['produto'] for p in (vendas + bonus)
                if p['produto'] not in todos_matched
                and mapa_atual.get(p['produto'], p['produto']) not in todos_matched
            ))
            if sem_slot:
                avisos.append(
                    f'⚠️ {len(sem_slot)} produto(s) sem cadastro (preencher manualmente): '
                    f'{"; ".join(sem_slot[:5])}'
                )

            ruins = [
                f"'{p['nome']}' → '{p['match_venda']}' ({int(p['score_venda']*100)}%)"
                for ps in agrupado.values() for p in ps
                if p.get('match_venda') and 0 < p.get('score_venda', 1) < 0.40
            ]
            if ruins:
                avisos.append(f'Matches incertos: {"; ".join(ruins[:3])}')

            est_map, prod_map = ler_mapa_linhas(service, sid, abas_mapa)
            msgs.append(f'Mapa: {len(est_map)} ESTOQUE, {len(prod_map)} PRODUÇÃO')

            est_updates = build_estoque_updates(agrupado, est_map, abas_mapa)
            batch.extend(est_updates)
            msgs.append(f'ESTOQUE col I: {len(est_updates)} células')

            prod_updates, prod_avisos = build_producao_updates(agrupado, prod_map, abas_mapa)
            batch.extend(prod_updates)
            msgs.append(f'PRODUÇÃO col C: {len(prod_updates)} células')
            avisos.extend(prod_avisos)

        # ── Caixas ────────────────────────────────────────────────────────
        if 'exportacao_caixas' in request.files:
            try:
                caixas = parse_caixas(request.files['exportacao_caixas'].read())
            except ValueError as e:
                return jsonify({'success': False, 'error': str(e)}), 400

            def op_norm(s):
                s = unicodedata.normalize('NFKD', str(s).upper())
                return ''.join(c for c in s if not unicodedata.combining(c))

            # Separação correta dos 3 tipos de operador do YUZER:
            # GARÇOM PIX   → Bloco 1 (L3:L32)   — garçons com máquina
            # Caixa PIX    → Bloco 2 (L36:L50)  — caixas volantes (PDV móvel)
            # CAIXA FIXO   → Bloco 3 (L54:L83)  — PDV fixo (alimentação, bar fixo, etc.)
            garcons   = [c for c in caixas if 'GARCOM' in op_norm(c['operacao'])]
            volantes  = [c for c in caixas if 'CAIXA'  in op_norm(c['operacao'])
                                           and 'FIXO'  not in op_norm(c['operacao'])]
            fixos     = [c for c in caixas if 'FIXO'   in op_norm(c['operacao'])]
            nome_cx   = aba(abas_mapa, 'FECHAMENTO CAIXAS')

            def to_rows(lista):
                return [[c['usuario'], c['serial'], c['total'],
                         c['dinheiro'], c['pix'], c['debito'], c['credito']]
                        for c in lista]

            # Bloco 1 — Garçons PIX (L3:L32)
            if garcons:
                rows = to_rows(garcons)
                if len(rows) > LIMITE_GARCONS:
                    avisos.append(
                        f'⚠️ {len(rows)} garçons — planilha tem {LIMITE_GARCONS} linhas '
                        f'(L3:L32). Adicione {len(rows)-LIMITE_GARCONS} linha(s) antes de L33.'
                    )
                batch.append({'range': f"{nome_cx}!B3:H{2+len(rows)}", 'values': rows})
                msgs.append(f'FECHAMENTO CAIXAS — Garçons PIX: {len(rows)}')

            # Bloco 2 — Caixas Volantes (L36:L50)
            if volantes:
                rows = to_rows(volantes)
                if len(rows) > LIMITE_VOLANTES:
                    avisos.append(
                        f'⚠️ {len(rows)} caixas volantes — planilha tem {LIMITE_VOLANTES} linhas '
                        f'(L36:L50). Adicione {len(rows)-LIMITE_VOLANTES} linha(s) antes de L51.'
                    )
                batch.append({'range': f"{nome_cx}!B36:H{35+len(rows)}", 'values': rows})
                msgs.append(f'FECHAMENTO CAIXAS — Caixas Volantes: {len(rows)}')

            # Bloco 3 — Caixas Fixos / PDV fixo (L54:L83)
            if fixos:
                rows = to_rows(fixos)
                if len(rows) > LIMITE_FIXOS:
                    avisos.append(
                        f'⚠️ {len(rows)} caixas fixos — planilha tem {LIMITE_FIXOS} linhas '
                        f'(L54:L83). Adicione {len(rows)-LIMITE_FIXOS} linha(s) antes de L84.'
                    )
                batch.append({'range': f"{nome_cx}!B54:H{53+len(rows)}", 'values': rows})
                msgs.append(f'FECHAMENTO CAIXAS — Caixas Fixos/PDV: {len(rows)}')

        # ── Painel → RESUMO col B ─────────────────────────────────────────
        if 'painel_de_vendas' in request.files:
            try:
                painel_data = parse_painel_vendas(request.files['painel_de_vendas'].read())
            except ValueError as e:
                return jsonify({'success': False, 'error': str(e)}), 400

            fp       = painel_data.get('formas_pagamento', {})
            total_fp = fp.get('CASH', 0) + fp.get('CREDIT_CARD', 0) + \
                       fp.get('DEBIT_CARD', 0) + fp.get('PIX', 0)
            nome_res = aba(abas_mapa, 'RESUMO')
            batch.append({
                'range': f'{nome_res}!B3:B9',
                'values': [
                    [0], [fp.get('CASH', 0)], [fp.get('CREDIT_CARD', 0)],
                    [fp.get('DEBIT_CARD', 0)], [fp.get('PIX', 0)], [0], [total_fp],
                ],
            })
            msgs.append('RESUMO col B preenchido')

            if agrupado:
                avisos.extend(validar_totais(agrupado, painel_data))

        # ── Enviar batch ──────────────────────────────────────────────────
        if batch:
            service.spreadsheets().values().batchUpdate(
                spreadsheetId=sid,
                body={
                    'valueInputOption': 'USER_ENTERED',
                    'data': [{'range': u['range'], 'values': u['values']} for u in batch],
                }
            ).execute()
            msgs.append(f'{len(batch)} operações gravadas')
        else:
            avisos.append('Nenhum dado para gravar — verifique os arquivos enviados')

        relatorio = gerar_relatorio_texto(agrupado, msgs, avisos, painel_data)

        return jsonify({
            'success':   True,
            'message':   'Dados enviados com sucesso!',
            'detalhes':  msgs,
            'avisos':    avisos,
            'relatorio': relatorio,
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error':   str(e),
            'trace':   traceback.format_exc(),
        }), 400


@app.route('/api/health')
def health():
    return jsonify({
        'status':              'ok',
        'app':                 'Prime Bar YUZER v5.2',
        'mapeamentos_globais': len(_mapeamento_global),
        'planilhas_mapeadas':  len(_mapeamento_store),
    })


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
